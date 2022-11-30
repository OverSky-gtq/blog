# coding=gbk
import openpyxl
wb1 = openpyxl.load_workbook('成绩.xlsx')
wb2 = openpyxl.load_workbook('信息.xlsx')
id = wb1['Sheet1']
info = wb2['Sheet1']
for i in range(5,id.max_row):
    target = str(id['B'+str(i)].value)
    for j in range(2,info.max_row):
        if target == str(info['C'+str(j)].value):
            print("success")
            id['D'+str(i)] = str(info['B'+str(j)].value)
            break
wb1.save("成功.xlsx")

